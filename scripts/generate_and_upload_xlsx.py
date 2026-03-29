import base64
import csv
import io
import json
import os
import re
import sys
from datetime import datetime, timedelta
from typing import Any, Dict

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

NOTION_VERSION = "2025-09-03"

DAYS_KR = ['월', '화', '수', '목', '금', '토', '일']


def load_event_payload(event_path: str) -> Dict[str, Any]:
    with open(event_path, "r", encoding="utf-8") as f:
        event = json.load(f)

    payload = event.get("client_payload") or event.get("inputs") or {}
    if not payload:
        raise RuntimeError("No payload found in GitHub event. Expected client_payload or inputs.")
    return payload


def unescape_csv_text(text: str) -> str:
    """Convert escaped JSON-string style CSV text back to raw CSV text."""
    return (
        text.replace("\\r\\n", "\n")
        .replace("\\n", "\n")
        .replace("\\r", "\n")
        .replace('\\\"', '"')
        .replace("\\\\", "\\")
    )


def fix_allday_schedule(value: str) -> str:
    """
    Make.com 날짜 형식에서 종일(all-day) 일정을 감지하여 처리.
    - 실제 포맷: YYYY.MM.DD(요일) HH:MM ~ YYYY.MM.DD(요일) HH:MM  (24시간)
    - 종일 감지: 시작/종료 시간이 모두 00:00
    - Google Calendar 종일 일정은 종료일이 실제 종료일 +1일(exclusive)이므로 -1일 조정
    - 시간 정보(00:00) 제거하여 날짜만 표시
    """
    if not isinstance(value, str):
        return value

    # ── YYYY.MM.DD(요일) HH:MM ~ YYYY.MM.DD(요일) HH:MM ──────────
    m = re.match(
        r'(\d{4})\.(\d{2})\.(\d{2})\([^)]+\)\s*(\d{2}):(\d{2})'
        r'\s*~\s*(\d{4})\.(\d{2})\.(\d{2})\([^)]+\)\s*(\d{2}):(\d{2})',
        value.strip()
    )
    if m:
        s_year, s_month, s_day = int(m.group(1)), int(m.group(2)), int(m.group(3))
        s_hh, s_mm = int(m.group(4)), int(m.group(5))
        e_year, e_month, e_day = int(m.group(6)), int(m.group(7)), int(m.group(8))
        e_hh, e_mm = int(m.group(9)), int(m.group(10))

        # 시작/종료 모두 00:00 → 종일 이벤트
        if s_hh == 0 and s_mm == 0 and e_hh == 0 and e_mm == 0:
            start_dt = datetime(s_year, s_month, s_day)
            end_dt   = datetime(e_year, e_month, e_day) - timedelta(days=1)  # 종료일 -1일 보정

            s_dow = DAYS_KR[start_dt.weekday()]
            if start_dt.date() == end_dt.date():
                # 1일짜리 종일 이벤트 → 날짜 하나만 표시
                return f"{s_year}.{s_month:02d}.{s_day:02d}({s_dow})"
            else:
                e_dow = DAYS_KR[end_dt.weekday()]
                return (
                    f"{s_year}.{s_month:02d}.{s_day:02d}({s_dow})"
                    f" ~ {end_dt.year}.{end_dt.month:02d}.{end_dt.day:02d}({e_dow})"
                )

    return value


def build_xlsx_from_csv_text(csv_text: str, delimiter: str = ",", report_title: str = "") -> bytes:
    from openpyxl.utils import get_column_letter

    reader = csv.reader(io.StringIO(csv_text), delimiter=delimiter)
    rows = list(reader)
    if not rows:
        raise RuntimeError("CSV text is empty.")

    # "A_구분", "B_일정" 등 순서 고정용 접두사(X_) 제거
    rows[0] = [
        col[2:] if len(col) > 2 and col[1] == "_" and col[0].isalpha() else col
        for col in rows[0]
    ]

    header = rows[0]
    data_rows = rows[1:]

    # ── 0. Sort: 구분 오름차순 → 일정 오름차순 ───────────────────
    try:
        gubun_idx = header.index("구분")
    except ValueError:
        gubun_idx = 0
    try:
        iljeong_idx = header.index("일정")
    except ValueError:
        iljeong_idx = 1
    try:
        bigo_idx = header.index("비고")
    except ValueError:
        bigo_idx = len(header) - 1

    data_rows.sort(key=lambda r: (
        r[gubun_idx] if gubun_idx < len(r) else "",
        r[iljeong_idx] if iljeong_idx < len(r) else "",
    ))

    # ── 종일 일정 날짜 보정 ────────────────────────────────────────
    for row in data_rows:
        if iljeong_idx < len(row):
            row[iljeong_idx] = fix_allday_schedule(row[iljeong_idx])

    rows = [header] + data_rows

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    num_cols = len(header)
    data_start_row = 1

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # ── 인쇄 설정: 가로 방향 + 자동맞춤 ─────────────────────────────
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # ── 1. Title row ──────────────────────────────────────────────
    # '콘텐츠IP전략팀'을 제목 앞에 항상 추가
    display_title = f"콘텐츠IP전략팀 {report_title}".strip() if report_title else "콘텐츠IP전략팀"

    title_cell = ws.cell(row=1, column=1, value=display_title)
    if num_cols > 1:
        ws.merge_cells(
            start_row=1, start_column=1,
            end_row=1,   end_column=num_cols,
        )
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36
    data_start_row = 2

    # ── 2. Write CSV rows (header + data) ────────────────────────
    for row_idx, row in enumerate(rows):
        for col_idx, value in enumerate(row):
            ws.cell(row=data_start_row + row_idx, column=col_idx + 1, value=value)

    # ── 3. Style header row (bold, size 14, center, light gray bg) ──
    header_row_idx = data_start_row
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    for cell in ws[header_row_idx]:
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        cell.fill = gray_fill
    ws.row_dimensions[header_row_idx].height = 24

    # ── 4. Style data rows + borders ─────────────────────────────
    for r_idx, row in enumerate(data_rows):
        excel_row = data_start_row + 1 + r_idx

        # 비고 필드에 "본부장" 포함 여부 확인
        bigo_value = row[bigo_idx] if bigo_idx < len(row) else ""
        is_honbujang = "본부장" in bigo_value

        for c_idx in range(num_cols):
            cell = ws.cell(row=excel_row, column=c_idx + 1)
            cell.border = border

            # 본부장 행: B~D열(2~4열) 노란색 배경 + 볼드
            if is_honbujang and 2 <= c_idx + 1 <= 4:
                cell.fill = yellow_fill
                cell.font = Font(bold=True, size=12)
            elif c_idx == gubun_idx:
                cell.font = Font(bold=True, size=12)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c_idx == bigo_idx:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="center")

    # ── 5. Merge consecutive same 구분 values ─────────────────────
    i = 0
    gubun_excel_col = gubun_idx + 1  # 1-based
    while i < len(data_rows):
        val = data_rows[i][gubun_idx] if gubun_idx < len(data_rows[i]) else ""
        j = i + 1
        while j < len(data_rows) and (data_rows[j][gubun_idx] if gubun_idx < len(data_rows[j]) else "") == val:
            j += 1
        if j - i > 1:
            start_excel_row = data_start_row + 1 + i
            end_excel_row = data_start_row + 1 + j - 1
            ws.merge_cells(
                start_row=start_excel_row, start_column=gubun_excel_col,
                end_row=end_excel_row,   end_column=gubun_excel_col,
            )
            anchor = ws.cell(row=start_excel_row, column=gubun_excel_col)
            anchor.font = Font(bold=True, size=12)
            anchor.alignment = Alignment(horizontal="center", vertical="center")
            anchor.border = border
        i = j

    # ── 6. Auto-fit column widths ─────────────────────────────────
    for col_idx in range(1, num_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row_cells in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row_cells:
                if cell.value is not None:
                    try:
                        v = str(cell.value).split("\n")[0]
                        if len(v) > max_len:
                            max_len = len(v)
                    except Exception:
                        pass

        if col_idx == bigo_idx + 1:
            # 비고: 최대 40자
            ws.column_dimensions[col_letter].width = min(max(8, max_len + 2), 40)
        elif col_idx == gubun_excel_col:
            # 구분: 내용에 맞게 자동 (최소 4, 최대 20)
            ws.column_dimensions[col_letter].width = min(max(4, max_len + 1), 20)
        else:
            ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 80)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def notion_headers(token: str, with_json: bool = False) -> Dict[str, str]:
    headers = {
        "Authorization": f"Bearer {token}",
        "Notion-Version": NOTION_VERSION,
    }
    if with_json:
        headers["Content-Type"] = "application/json"
    return headers


def upload_xlsx_to_notion(token: str, file_bytes: bytes, filename: str) -> str:
    create_resp = requests.post(
        "https://api.notion.com/v1/file_uploads",
        headers=notion_headers(token, with_json=True),
        json={
            "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "filename": filename,
        },
        timeout=30,
    )
    create_resp.raise_for_status()
    create_obj = create_resp.json()

    file_upload_id = create_obj["id"]
    upload_url = create_obj["upload_url"]

    send_resp = requests.post(
        upload_url,
        headers=notion_headers(token),
        files={"file": (filename, file_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        timeout=60,
    )
    send_resp.raise_for_status()
    return file_upload_id


def attach_file_to_page(token: str, page_id: str, file_upload_id: str, filename: str, file_property_name: str) -> None:
    body = {
        "properties": {
            file_property_name: {
                "files": [
                    {
                        "type": "file_upload",
                        "file_upload": {"id": file_upload_id},
                        "name": filename,
                    }
                ]
            }
        }
    }

    patch_resp = requests.patch(
        f"https://api.notion.com/v1/pages/{page_id}",
        headers=notion_headers(token, with_json=True),
        json=body,
        timeout=30,
    )
    patch_resp.raise_for_status()


def main() -> int:
    notion_token = os.getenv("NOTION_TOKEN", "").strip()
    if not notion_token:
        raise RuntimeError("NOTION_TOKEN secret is required.")

    event_path = os.getenv("GITHUB_EVENT_PATH", "")
    if not event_path:
        raise RuntimeError("GITHUB_EVENT_PATH is missing.")

    payload = load_event_payload(event_path)
    csv_text = payload.get("csv_text", "")
    csv_b64 = payload.get("csv_b64", "")
    page_id = payload.get("notion_page_id", "")
    filename = payload.get("report_name", "weekly_report.xlsx")
    file_property_name = payload.get("file_property_name", "파일과 미디어")
    delimiter = payload.get("delimiter", ",")
    report_title = payload.get("report_title", "")

    if not csv_text and not csv_b64:
        raise RuntimeError("payload.csv_text or payload.csv_b64 is required.")
    if not page_id:
        raise RuntimeError("payload.notion_page_id is required.")

    if csv_b64:
        csv_text = base64.b64decode(csv_b64).decode("utf-8", errors="replace")
    else:
        csv_text = unescape_csv_text(csv_text)

    if delimiter == "tab":
        delimiter = "\t"

    xlsx_bytes = build_xlsx_from_csv_text(csv_text, delimiter=delimiter, report_title=report_title)
    file_upload_id = upload_xlsx_to_notion(notion_token, xlsx_bytes, filename)
    attach_file_to_page(notion_token, page_id, file_upload_id, filename, file_property_name)

    print("OK")
    print(json.dumps({"page_id": page_id, "filename": filename, "file_upload_id": file_upload_id}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr)
        raise
